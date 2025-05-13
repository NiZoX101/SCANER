import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from heuristics import detailed_extract_suspicious_lines
import re
import ast
from datetime import datetime
import os

def generate_detailed_report(results, scores, output_path, check_types, report_mode="overall"):
    """
    Generate a detailed report in Excel format.
    
    Parameters:
    - results: Dict from scanner_core.analyze_code with findings for each file.
    - scores: Dict with risk scores for each file.
    - output_path: Path to save the Excel file(s).
    - check_types: List of check types performed (e.g., ["ast", "regex", "yara", "bandit", "heuristics"]).
    - report_mode: "overall" for a single report, "separate" for reports by check type.
    """
    if report_mode == "overall":
        generate_overall_report(results, scores, output_path, check_types)
    elif report_mode == "separate":
        generate_separate_reports(results, scores, output_path, check_types)

def generate_overall_report(results, scores, output_path, check_types):
    """Generate a single Excel report combining all check types."""
    wb = openpyxl.Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    # Define styles
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="487eb0", end_color="487eb0", fill_type="solid")
    cell_border = Border(left=Side(style="thin"), right=Side(style="thin"), 
                         top=Side(style="thin"), bottom=Side(style="thin"))
    center_align = Alignment(horizontal="center", vertical="center")
    
    # Summary Sheet
    ws_summary.append(["Advanced Code Security Scanner - Summary"])
    ws_summary.merge_cells("A1:E1")
    ws_summary["A1"].font = Font(bold=True, size=14)
    ws_summary["A1"].alignment = center_align
    
    total_files = len(scores)
    avg_score = sum(scores.values()) / total_files if total_files > 0 else 0
    high_risk = sum(1 for s in scores.values() if s >= 70)
    
    ws_summary.append(["Total Files Scanned", total_files])
    ws_summary.append(["Average Risk Score", f"{avg_score:.2f}"])
    ws_summary.append(["High Risk Files (Score >= 70)", high_risk])
    ws_summary.append(["Check Types Used", ", ".join(check_types)])
    ws_summary.append(["Generated On", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    
    for row in ws_summary["A2:E6"]:
        for cell in row:
            cell.border = cell_border
            if cell.row == 2:
                cell.alignment = center_align
    
    # Findings Sheet
    ws_findings = wb.create_sheet("Detailed Findings")
    headers = ["File", "Risk Score", "Check Type", "Finding", "Line Number", "Code Snippet", "Description", "Advice"]
    ws_findings.append(headers)
    
    for col, header in enumerate(headers, 1):
        cell = ws_findings.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = cell_border
    
    row_idx = 2
    for filename, patterns in results.items():
        score = scores.get(filename, 0)
        
        # AST Checks
        if "ast" in check_types:
            ast_checks = {
                "eval_usage": ("Use of eval()", "Avoid eval; use safer alternatives like ast.literal_eval."),
                "exec_usage": ("Use of exec()", "Avoid exec; restrict dynamic code execution."),
                "subprocess_usage": ("Subprocess calls", "Validate subprocess arguments to prevent injection."),
                "os_system": ("os.system call", "Use subprocess with validated arguments instead."),
                "pickle_loads": ("pickle.loads usage", "Use safer serialization formats like JSON."),
                "subprocess_shell_true": ("Subprocess with shell=True", "Avoid shell=True; use shell=False with validated args."),
                "insecure_serialization": ("Insecure serialization (pickle/marshal/shelve)", "Use JSON or other safe formats."),
                "input_usage": ("Use of input()", "Validate user input to prevent injection."),
                "dynamic_import": ("Dynamic import (__import__)", "Validate module sources before importing.")
            }
            for key, (desc, advice) in ast_checks.items():
                if patterns.get(key, False):
                    lines = find_line_numbers(filename, patterns, key)
                    for line_num, snippet in lines:
                        ws_findings.append([
                            filename, score, "AST", key, line_num, snippet, desc, advice
                        ])
                        for col in range(1, 9):
                            ws_findings.cell(row=row_idx, column=col).border = cell_border
                        row_idx += 1
        
        # Regex Checks
        if "regex" in check_types:
            regex_checks = {
                "ip_blocking": ("IP blocking pattern", "Review blacklist logic for potential abuse."),
                "spam_subscription": ("Spam subscription pattern", "Verify email subscription logic."),
                "dangerous_links": ("Dangerous file links", "Avoid links to executable files."),
                "malicious_download": ("Suspicious download", "Validate download sources."),
                "user_data_exfiltration": ("Data exfiltration attempt", "Secure file operations."),
                "registry_access": ("Windows registry access", "Verify necessity of registry operations."),
                "base64_decode_eval": ("Base64 decode with eval", "Avoid decoding and executing dynamic code."),
                "suspicious_dynamic_exec": ("Dynamic execution pattern", "Restrict dynamic code execution."),
                "dangerous_imports": ("Dangerous imports", "Review use of socket, ctypes, etc."),
                "obfuscated_code": ("Obfuscated code", "Clarify code to avoid hiding malicious intent."),
                "suspicious_file_ops": ("Suspicious file operations", "Validate file write operations."),
                "dynamic_module_loading": ("Dynamic module loading", "Validate imported modules."),
                "suspicious_getattr": ("Suspicious getattr usage", "Ensure getattr is used safely."),
                "hardcoded_credentials": ("Hardcoded credentials", "Store credentials in environment variables."),
                "insecure_protocol": ("Use of HTTP", "Use HTTPS for secure communication."),
                "sql_injection": ("Potential SQL injection", "Use parameterized queries."),
                "reverse_shell": ("Reverse shell pattern", "Remove or secure network connections."),
                "weak_encryption": ("Weak encryption (MD5/SHA1)", "Use stronger algorithms like SHA-256."),
                "malicious_comments": ("Malicious comments", "Remove suspicious comments."),
                "hidden_process": ("Hidden process creation", "Avoid hiding processes."),
                "camera_access": ("Camera access", "Verify necessity and user consent."),
                "microphone_access": ("Microphone access", "Verify necessity and user consent."),
                "dropper_code": ("Dropper code pattern", "Remove or secure dynamic code execution."),
                "code_injection": ("Code injection attempt", "Secure memory operations."),
                "file_permission_changes": ("File permission changes", "Validate permission modifications."),
                "in_memory_execution": ("In-memory code execution", "Avoid executing decoded code."),
                "env_variable_usage": ("Environment variable access", "Validate variable usage.")
            }
            for key, (desc, advice) in regex_checks.items():
                if patterns.get(key, False):
                    lines = find_line_numbers(filename, patterns, key)
                    for line_num, snippet in lines:
                        ws_findings.append([
                            filename, score, "Regex", key, line_num, snippet, desc, advice
                        ])
                        for col in range(1, 9):
                            ws_findings.cell(row=row_idx, column=col).border = cell_border
                        row_idx += 1
        
        # YARA Check
        if "yara" in check_types and patterns.get("yara_match", False):
            ws_findings.append([
                filename, score, "YARA", "yara_match", "-", "N/A", 
                "YARA rule match detected", "Review YARA rule details and code context."
            ])
            for col in range(1, 9):
                ws_findings.cell(row=row_idx, column=col).border = cell_border
            row_idx += 1
        
        # Bandit Checks
        if "bandit" in check_types:
            bandit_checks = {
                "assert_used": ("Use of assert", "Avoid assert in production code."),
                "exec_used": ("Use of exec", "Restrict dynamic code execution."),
                "bad_file_perms": ("Insecure file permissions", "Use restrictive permissions (e.g., 600)."),
                "bind_all_interfaces": ("Binding to all interfaces", "Bind to specific interfaces."),
                "hardcoded_password": ("Hardcoded password", "Use environment variables."),
                "hardcoded_tmp": ("Hardcoded temporary directory", "Use tempfile module."),
                "hardcoded_ssl_cert": ("Hardcoded SSL certificate", "Store certificates securely."),
                "aws_keys": ("Hardcoded AWS keys", "Use AWS Secrets Manager."),
                "pickle_used": ("Use of pickle", "Use JSON or other safe formats."),
                "marshal_used": ("Use of marshal", "Avoid marshal for serialization."),
                "md5_used": ("Use of MD5", "Use SHA-256 or stronger algorithms."),
                "cgi_used": ("Use of cgi module", "Avoid cgi; use modern frameworks."),
                "ftplib_used": ("Use of ftplib", "Use secure alternatives like sftp."),
                "mktemp_used": ("Use of mktemp", "Use tempfile.TemporaryFile."),
                "eval_used": ("Use of eval", "Use safer alternatives."),
                "mark_safe_used": ("Django mark_safe usage", "Validate input before marking safe."),
                "httpsconnection_used": ("Insecure HTTPS connection", "Use modern TLS settings."),
                "urlopen_used": ("urllib.urlopen usage", "Use requests with timeout and validation."),
                "random_used": ("Use of random module", "Use secrets module for cryptographic tasks."),
                "telnetlib_used": ("Use of telnetlib", "Use SSH or secure protocols."),
                "cElementTree_used": ("Use of cElementTree", "Use defusedxml for XML parsing."),
                "paramiko_insecure": ("Insecure Paramiko usage", "Set secure host key policies."),
                "ssl_bad_version": ("Insecure SSL/TLS version", "Use TLS 1.2 or higher."),
                "ssl_bad_defaults": ("Insecure SSL defaults", "Configure secure SSL settings."),
                "paramiko_exec_command": ("Paramiko exec_command", "Validate commands before execution."),
                "subprocess_shell_true": ("Subprocess with shell=True", "Use shell=False with validated args."),
                "subprocess_without_shell": ("Subprocess without shell", "Ensure proper argument validation."),
                "other_shell_true": ("Other shell=True usage", "Avoid shell=True in custom functions."),
                "partial_path_process": ("Partial path in process", "Use full paths for executables."),
                "shell_process": ("Shell process execution", "Use subprocess with validation."),
                "hardcoded_sql": ("Hardcoded SQL queries", "Use parameterized queries."),
                "wildcard_injection": ("Wildcard injection in commands", "Avoid wildcards in commands."),
                "django_extra_used": ("Django extra() usage", "Use ORM methods instead."),
                "django_rawsql_used": ("Django raw SQL usage", "Use parameterized queries."),
                "logging_insecure": ("Insecure logging config", "Avoid SocketHandler in logging."),
                "jinja2_autoescape_off": ("Jinja2 autoescape off", "Enable autoescape for security.")
            }
            for key, (desc, advice) in bandit_checks.items():
                if patterns.get(key, False):
                    lines = find_line_numbers(filename, patterns, key)
                    for line_num, snippet in lines:
                        ws_findings.append([
                            filename, score, "Bandit", key, line_num, snippet, desc, advice
                        ])
                        for col in range(1, 9):
                            ws_findings.cell(row=row_idx, column=col).border = cell_border
                        row_idx += 1
        
        # Heuristic Checks
        if "heuristics" in check_types:
            # ML Suspicion
            ml_score = patterns.get("ml_suspicion", 0)
            if ml_score > 0:
                ws_findings.append([
                    filename, score, "Heuristics", "ml_suspicion", "-", "N/A",
                    f"ML-based suspicion score: {ml_score}", "Review code for suspicious patterns."
                ])
                for col in range(1, 9):
                    ws_findings.cell(row=row_idx, column=col).border = cell_border
                row_idx += 1
            
            # Detailed Lines (from heuristics.detailed_extract_suspicious_lines)
            detailed_lines = patterns.get("detailed_lines", [])
            for line in detailed_lines:
                # Parse line format: "Line X: <code> [Причина: <desc>] [Совет: <advice>]"
                match = re.match(r"Line (\d+): (.*?) \[Причина: (.*?)\] \[Совет: (.*?)\]", line)
                if match:
                    line_num, snippet, desc, advice = match.groups()
                    ws_findings.append([
                        filename, score, "Heuristics", "suspicious_line", line_num, snippet, desc, advice
                    ])
                    for col in range(1, 9):
                        ws_findings.cell(row=row_idx, column=col).border = cell_border
                    row_idx += 1
            
            # API Keys
            api_keys = patterns.get("api_keys", {})
            for key, info in api_keys.items():
                if info.get("found", False):
                    ws_findings.append([
                        filename, score, "Heuristics", key, "-", info.get("match", "N/A"),
                        info.get("description", ""), info.get("advice", "")
                    ])
                    for col in range(1, 9):
                        ws_findings.cell(row=row_idx, column=col).border = cell_border
                    row_idx += 1
            
            # Encryption
            encryption = patterns.get("encryption", {})
            for key, info in encryption.items():
                if info.get("found", False):
                    lines = find_line_numbers(filename, patterns, key)
                    for line_num, snippet in lines:
                        ws_findings.append([
                            filename, score, "Heuristics", key, line_num, snippet,
                            info.get("description", ""), info.get("advice", "")
                        ])
                        for col in range(1, 9):
                            ws_findings.cell(row=row_idx, column=col).border = cell_border
                        row_idx += 1
            
            # Network
            network = patterns.get("network", {})
            for key, info in network.items():
                if info.get("found", False):
                    lines = find_line_numbers(filename, patterns, key)
                    for line_num, snippet in lines:
                        ws_findings.append([
                            filename, score, "Heuristics", key, line_num, snippet,
                            info.get("description", ""), info.get("advice", "")
                        ])
                        for col in range(1, 9):
                            ws_findings.cell(row=row_idx, column=col).border = cell_border
                        row_idx += 1
            
            # Hardware
            hardware = patterns.get("hardware", {})
            for key, info in hardware.items():
                if info.get("found", False):
                    lines = find_line_numbers(filename, patterns, key)
                    for line_num, snippet in lines:
                        ws_findings.append([
                            filename, score, "Heuristics", key, line_num, snippet,
                            info.get("description", ""), info.get("advice", "")
                        ])
                        for col in range(1, 9):
                            ws_findings.cell(row=row_idx, column=col).border = cell_border
                        row_idx += 1
            
            # Shell Injection
            shell_injection = patterns.get("shell_injection", {})
            if shell_injection.get("found", False):
                lines = find_line_numbers(filename, patterns, "shell_injection")
                for line_num, snippet in lines:
                    ws_findings.append([
                        filename, score, "Heuristics", "shell_injection", line_num, snippet,
                        shell_injection.get("description", ""), shell_injection.get("advice", "")
                    ])
                    for col in range(1, 9):
                        ws_findings.cell(row=row_idx, column=col).border = cell_border
                    row_idx += 1
            
            # SQL Injection
            sql_injection = patterns.get("sql_injection", {})
            if sql_injection.get("found", False):
                lines = find_line_numbers(filename, patterns, "sql_injection")
                for line_num, snippet in lines:
                    ws_findings.append([
                        filename, score, "Heuristics", "sql_injection", line_num, snippet,
                        sql_injection.get("description", ""), sql_injection.get("advice", "")
                    ])
                    for col in range(1, 9):
                        ws_findings.cell(row=row_idx, column=col).border = cell_border
                    row_idx += 1
    
    # Adjust column widths
    for col in range(1, 9):
        max_length = 0
        for row in ws_findings[f"{get_column_letter(col)}"]:
            try:
                if len(str(row.value)) > max_length:
                    max_length = len(str(row.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_findings.column_dimensions[get_column_letter(col)].width = adjusted_width
    
    wb.save(output_path)

def generate_separate_reports(results, scores, output_path, check_types):
    """Generate separate Excel reports for each check type."""
    base_path, ext = os.path.splitext(output_path)
    for check_type in check_types:
        wb = openpyxl.Workbook()
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        # Define styles
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="487eb0", end_color="487eb0", fill_type="solid")
        cell_border = Border(left=Side(style="thin"), right=Side(style="thin"), 
                             top=Side(style="thin"), bottom=Side(style="thin"))
        center_align = Alignment(horizontal="center", vertical="center")
        
        # Summary Sheet
        ws_summary.append([f"Advanced Code Security Scanner - {check_type.capitalize()} Report"])
        ws_summary.merge_cells("A1:E1")
        ws_summary["A1"].font = Font(bold=True, size=14)
        ws_summary["A1"].alignment = center_align
        
        total_files = len(scores)
        avg_score = sum(scores.values()) / total_files if total_files > 0 else 0
        high_risk = sum(1 for s in scores.values() if s >= 70)
        
        ws_summary.append(["Total Files Scanned", total_files])
        ws_summary.append(["Average Risk Score", f"{avg_score:.2f}"])
        ws_summary.append(["High Risk Files (Score >= 70)", high_risk])
        ws_summary.append(["Check Type", check_type])
        ws_summary.append(["Generated On", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
        
        for row in ws_summary["A2:E6"]:
            for cell in row:
                cell.border = cell_border
                if cell.row == 2:
                    cell.alignment = center_align
        
        # Findings Sheet
        ws_findings = wb.create_sheet("Detailed Findings")
        headers = ["File", "Risk Score", "Finding", "Line Number", "Code Snippet", "Description", "Advice"]
        ws_findings.append(headers)
        
        for col, header in enumerate(headers, 1):
            cell = ws_findings.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = cell_border
        
        row_idx = 2
        for filename, patterns in results.items():
            score = scores.get(filename, 0)
            
            if check_type == "ast":
                ast_checks = {
                    "eval_usage": ("Use of eval()", "Avoid eval; use safer alternatives like ast.literal_eval."),
                    "exec_usage": ("Use of exec()", "Avoid exec; restrict dynamic code execution."),
                    "subprocess_usage": ("Subprocess calls", "Validate subprocess arguments to prevent injection."),
                    "os_system": ("os.system call", "Use subprocess with validated arguments instead."),
                    "pickle_loads": ("pickle.loads usage", "Use safer serialization formats like JSON."),
                    "subprocess_shell_true": ("Subprocess with shell=True", "Avoid shell=True; use shell=False with validated args."),
                    "insecure_serialization": ("Insecure serialization (pickle/marshal/shelve)", "Use JSON or other safe formats."),
                    "input_usage": ("Use of input()", "Validate user input to prevent injection."),
                    "dynamic_import": ("Dynamic import (__import__)", "Validate module sources before importing.")
                }
                for key, (desc, advice) in ast_checks.items():
                    if patterns.get(key, False):
                        lines = find_line_numbers(filename, patterns, key)
                        for line_num, snippet in lines:
                            ws_findings.append([
                                filename, score, key, line_num, snippet, desc, advice
                            ])
                            for col in range(1, 8):
                                ws_findings.cell(row=row_idx, column=col).border = cell_border
                            row_idx += 1
            
            elif check_type == "regex":
                regex_checks = {
                    "ip_blocking": ("IP blocking pattern", "Review blacklist logic for potential abuse."),
                    "spam_subscription": ("Spam subscription pattern", "Verify email subscription logic."),
                    "dangerous_links": ("Dangerous file links", "Avoid links to executable files."),
                    "malicious_download": ("Suspicious download", "Validate download sources."),
                    "user_data_exfiltration": ("Data exfiltration attempt", "Secure file operations."),
                    "registry_access": ("Windows registry access", "Verify necessity of registry operations."),
                    "base64_decode_eval": ("Base64 decode with eval", "Avoid decoding and executing dynamic code."),
                    "suspicious_dynamic_exec": ("Dynamic execution pattern", "Restrict dynamic code execution."),
                    "dangerous_imports": ("Dangerous imports", "Review use of socket, ctypes, etc."),
                    "obfuscated_code": ("Obfuscated code", "Clarify code to avoid hiding malicious intent."),
                    "suspicious_file_ops": ("Suspicious file operations", "Validate file write operations."),
                    "dynamic_module_loading": ("Dynamic module loading", "Validate imported modules."),
                    "suspicious_getattr": ("Suspicious getattr usage", "Ensure getattr is used safely."),
                    "hardcoded_credentials": ("Hardcoded credentials", "Store credentials in environment variables."),
                    "insecure_protocol": ("Use of HTTP", "Use HTTPS for secure communication."),
                    "sql_injection": ("Potential SQL injection", "Use parameterized queries."),
                    "reverse_shell": ("Reverse shell pattern", "Remove or secure network connections."),
                    "weak_encryption": ("Weak encryption (MD5/SHA1)", "Use stronger algorithms like SHA-256."),
                    "malicious_comments": ("Malicious comments", "Remove suspicious comments."),
                    "hidden_process": ("Hidden process creation", "Avoid hiding processes."),
                    "camera_access": ("Camera access", "Verify necessity and user consent."),
                    "microphone_access": ("Microphone access", "Verify necessity and user consent."),
                    "dropper_code": ("Dropper code pattern", "Remove or secure dynamic code execution."),
                    "code_injection": ("Code injection attempt", "Secure memory operations."),
                    "file_permission_changes": ("File permission changes", "Validate permission modifications."),
                    "in_memory_execution": ("In-memory code execution", "Avoid executing decoded code."),
                    "env_variable_usage": ("Environment variable access", "Validate variable usage.")
                }
                for key, (desc, advice) in regex_checks.items():
                    if patterns.get(key, False):
                        lines = find_line_numbers(filename, patterns, key)
                        for line_num, snippet in lines:
                            ws_findings.append([
                                filename, score, key, line_num, snippet, desc, advice
                            ])
                            for col in range(1, 8):
                                ws_findings.cell(row=row_idx, column=col).border = cell_border
                            row_idx += 1
            
            elif check_type == "yara":
                if patterns.get("yara_match", False):
                    ws_findings.append([
                        filename, score, "yara_match", "-", "N/A", 
                        "YARA rule match detected", "Review YARA rule details and code context."
                    ])
                    for col in range(1, 8):
                        ws_findings.cell(row=row_idx, column=col).border = cell_border
                    row_idx += 1
            
            elif check_type == "bandit":
                bandit_checks = {
                    "assert_used": ("Use of assert", "Avoid assert in production code."),
                    "exec_used": ("Use of exec", "Restrict dynamic code execution."),
                    "bad_file_perms": ("Insecure file permissions", "Use restrictive permissions (e.g., 600)."),
                    "bind_all_interfaces": ("Binding to all interfaces", "Bind to specific interfaces."),
                    "hardcoded_password": ("Hardcoded password", "Use environment variables."),
                    "hardcoded_tmp": ("Hardcoded temporary directory", "Use tempfile module."),
                    "hardcoded_ssl_cert": ("Hardcoded SSL certificate", "Store certificates securely."),
                    "aws_keys": ("Hardcoded AWS keys", "Use AWS Secrets Manager."),
                    "pickle_used": ("Use of pickle", "Use JSON or other safe formats."),
                    "marshal_used": ("Use of marshal", "Avoid marshal for serialization."),
                    "md5_used": ("Use of MD5", "Use SHA-256 or stronger algorithms."),
                    "cgi_used": ("Use of cgi module", "Avoid cgi; use modern frameworks."),
                    "ftplib_used": ("Use of ftplib", "Use secure alternatives like sftp."),
                    "mktemp_used": ("Use of mktemp", "Use tempfile.TemporaryFile."),
                    "eval_used": ("Use of eval", "Use safer alternatives."),
                    "mark_safe_used": ("Django mark_safe usage", "Validate input before marking safe."),
                    "httpsconnection_used": ("Insecure HTTPS connection", "Use modern TLS settings."),
                    "urlopen_used": ("urllib.urlopen usage", "Use requests with timeout and validation."),
                    "random_used": ("Use of random module", "Use secrets module for cryptographic tasks."),
                    "telnetlib_used": ("Use of telnetlib", "Use SSH or secure protocols."),
                    "cElementTree_used": ("Use of cElementTree", "Use defusedxml for XML parsing."),
                    "paramiko_insecure": ("Insecure Paramiko usage", "Set secure host key policies."),
                    "ssl_bad_version": ("Insecure SSL/TLS version", "Use TLS 1.2 or higher."),
                    "ssl_bad_defaults": ("Insecure SSL defaults", "Configure secure SSL settings."),
                    "paramiko_exec_command": ("Paramiko exec_command", "Validate commands before execution."),
                    "subprocess_shell_true": ("Subprocess with shell=True", "Use shell=False with validated args."),
                    "subprocess_without_shell": ("Subprocess without shell", "Ensure proper argument validation."),
                    "other_shell_true": ("Other shell=True usage", "Avoid shell=True in custom functions."),
                    "partial_path_process": ("Partial path in process", "Use full paths for executables."),
                    "shell_process": ("Shell process execution", "Use subprocess with validation."),
                    "hardcoded_sql": ("Hardcoded SQL queries", "Use parameterized queries."),
                    "wildcard_injection": ("Wildcard injection in commands", "Avoid wildcards in commands."),
                    "django_extra_used": ("Django extra() usage", "Use ORM methods instead."),
                    "django_rawsql_used": ("Django raw SQL usage", "Use parameterized queries."),
                    "logging_insecure": ("Insecure logging config", "Avoid SocketHandler in logging."),
                    "jinja2_autoescape_off": ("Jinja2 autoescape off", "Enable autoescape for security.")
                }
                for key, (desc, advice) in bandit_checks.items():
                    if patterns.get(key, False):
                        lines = find_line_numbers(filename, patterns, key)
                        for line_num, snippet in lines:
                            ws_findings.append([
                                filename, score, key, line_num, snippet, desc, advice
                            ])
                            for col in range(1, 8):
                                ws_findings.cell(row=row_idx, column=col).border = cell_border
                            row_idx += 1
            
            elif check_type == "heuristics":
                # ML Suspicion
                ml_score = patterns.get("ml_suspicion", 0)
                if ml_score > 0:
                    ws_findings.append([
                        filename, score, "ml_suspicion", "-", "N/A",
                        f"ML-based suspicion score: {ml_score}", "Review code for suspicious patterns."
                    ])
                    for col in range(1, 8):
                        ws_findings.cell(row=row_idx, column=col).border = cell_border
                    row_idx += 1
                
                # Detailed Lines
                detailed_lines = patterns.get("detailed_lines", [])
                for line in detailed_lines:
                    match = re.match(r"Line (\d+): (.*?) \[Причина: (.*?)\] \[Совет: (.*?)\]", line)
                    if match:
                        line_num, snippet, desc, advice = match.groups()
                        ws_findings.append([
                            filename, score, "suspicious_line", line_num, snippet, desc, advice
                        ])
                        for col in range(1, 8):
                            ws_findings.cell(row=row_idx, column=col).border = cell_border
                        row_idx += 1
                
                # API Keys
                api_keys = patterns.get("api_keys", {})
                for key, info in api_keys.items():
                    if info.get("found", False):
                        ws_findings.append([
                            filename, score, key, "-", info.get("match", "N/A"),
                            info.get("description", ""), info.get("advice", "")
                        ])
                        for col in range(1, 8):
                            ws_findings.cell(row=row_idx, column=col).border = cell_border
                        row_idx += 1
                
                # Encryption
                encryption = patterns.get("encryption", {})
                for key, info in encryption.items():
                    if info.get("found", False):
                        lines = find_line_numbers(filename, patterns, key)
                        for line_num, snippet in lines:
                            ws_findings.append([
                                filename, score, key, line_num, snippet,
                                info.get("description", ""), info.get("advice", "")
                            ])
                            for col in range(1, 8):
                                ws_findings.cell(row=row_idx, column=col).border = cell_border
                            row_idx += 1
                
                # Network
                network = patterns.get("network", {})
                for key, info in network.items():
                    if info.get("found", False):
                        lines = find_line_numbers(filename, patterns, key)
                        for line_num, snippet in lines:
                            ws_findings.append([
                                filename, score, key, line_num, snippet,
                                info.get("description", ""), info.get("advice", "")
                            ])
                            for col in range(1, 8):
                                ws_findings.cell(row=row_idx, column=col).border = cell_border
                            row_idx += 1
                
                # Hardware
                hardware = patterns.get("hardware", {})
                for key, info in hardware.items():
                    if info.get("found", False):
                        lines = find_line_numbers(filename, patterns, key)
                        for line_num, snippet in lines:
                            ws_findings.append([
                                filename, score, key, line_num, snippet,
                                info.get("description", ""), info.get("advice", "")
                            ])
                            for col in range(1, 8):
                                ws_findings.cell(row=row_idx, column=col).border = cell_border
                            row_idx += 1
                
                # Shell Injection
                shell_injection = patterns.get("shell_injection", {})
                if shell_injection.get("found", False):
                    lines = find_line_numbers(filename, patterns, "shell_injection")
                    for line_num, snippet in lines:
                        ws_findings.append([
                            filename, score, "shell_injection", line_num, snippet,
                            shell_injection.get("description", ""), shell_injection.get("advice", "")
                        ])
                        for col in range(1, 8):
                            ws_findings.cell(row=row_idx, column=col).border = cell_border
                        row_idx += 1
                
                # SQL Injection
                sql_injection = patterns.get("sql_injection", {})
                if sql_injection.get("found", False):
                    lines = find_line_numbers(filename, patterns, "sql_injection")
                    for line_num, snippet in lines:
                        ws_findings.append([
                            filename, score, "sql_injection", line_num, snippet,
                            sql_injection.get("description", ""), sql_injection.get("advice", "")
                        ])
                        for col in range(1, 8):
                            ws_findings.cell(row=row_idx, column=col).border = cell_border
                        row_idx += 1
        
        # Adjust column widths
        for col in range(1, 8):
            max_length = 0
            for row in ws_findings[f"{get_column_letter(col)}"]:
                try:
                    if len(str(row.value)) > max_length:
                        max_length = len(str(row.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_findings.column_dimensions[get_column_letter(col)].width = adjusted_width
        
        wb.save(f"{base_path}_{check_type}{ext}")

def find_line_numbers(filename, patterns, key):
    """
    Find line numbers and code snippets for a given pattern.
    Returns list of (line_number, snippet) tuples.
    """
    code = patterns.get("_code", "")  # Assume code is stored in results
    if not code:
        return [("-", "N/A")]
    
    lines = code.splitlines()
    findings = []
    
    # For heuristics.detailed_extract_suspicious_lines, use existing line numbers
    if key == "suspicious_line" and "detailed_lines" in patterns:
        for line in patterns["detailed_lines"]:
            match = re.match(r"Line (\d+): (.*?) \[Причина: (.*?)\] \[Совет: (.*?)\]", line)
            if match:
                line_num, snippet, _, _ = match.groups()
                findings.append((line_num, snippet))
        return findings
    
    # For other patterns, search for matches
    pattern_map = {
        "eval_usage": r"eval\(",
        "exec_usage": r"exec\(",
        "subprocess_usage": r"subprocess\.(Popen|call|run)\(",
        "os_system": r"os\.system\(",
        "pickle_loads": r"pickle\.loads\(",
        "subprocess_shell_true": r"subprocess\.(Popen|call|run)\([^)]*shell\s*=\s*True",
        "insecure_serialization": r"(pickle|marshal|shelve)\.(loads|load)\(",
        "input_usage": r"input\(",
        "dynamic_import": r"__import__\(",
        "ip_blocking": r"if .*in.*blacklist",
        "spam_subscription": r"subscribe.*mail",
        "dangerous_links": r"http[s]?://.*(exe|bat|js|vbs|sh|dll)",
        "malicious_download": r"requests\.get\s*\(.*\b(url|path)\b",
        "user_data_exfiltration": r"open\s*\(.*\.(txt|csv|log)\)",
        "registry_access": r"winreg",
        "base64_decode_eval": r"base64\.b64decode\s*\([^)]*\)\s*.*eval\s*\(",
        "suspicious_dynamic_exec": r"(compile|exec)\s*\(",
        "dangerous_imports": r"(socket|ctypes|pickle|marshal|Crypto)",
        "obfuscated_code": r"\\x[0-9A-Fa-f]{2}",
        "suspicious_file_ops": r"open\s*\(.*[wa]\)",
        "dynamic_module_loading": r"(importlib\.import_module|__import__)",
        "suspicious_getattr": r"getattr\s*\(.*['\"]",
        "hardcoded_credentials": r"(password|secret|api_key)\s*=\s*['\"][^'\"]+['\"]",
        "insecure_protocol": r"http://",
        "sql_injection": r"execute\s*\(.*%.*\)",
        "reverse_shell": r"socket\.(socket|create_connection)\s*\(",
        "weak_encryption": r"cryptography\.(md5|sha1)",
        "malicious_comments": r"(backdoor|malware|exploit|keylogger|ransom)",
        "hidden_process": r"CREATE_NO_WINDOW|SW_HIDE",
        "camera_access": r"cv2\.VideoCapture|pygame\.camera",
        "microphone_access": r"sounddevice\.rec|pyaudio\.PyAudio",
        "dropper_code": r"requests\.get\(.*\)\.content.*exec\(|urllib\.request\.urlopen\(.*\)\.read\(\)",
        "code_injection": r"ctypes\.windll|WriteProcessMemory",
        "file_permission_changes": r"os\.chmod|os\.chown",
        "in_memory_execution": r"exec\(compile\(|eval\(compile\(|exec\(.*decode\(['\"]base64",
        "env_variable_usage": r"os\.getenv|os\.environ\.get",
        "assert_used": r"\bassert\b",
        "exec_used": r"exec\(",
        "bad_file_perms": r"os\.chmod\s*\(.*0o?777|0o?666",
        "bind_all_interfaces": r"\.bind\s*\(\s*['\"](0\.0\.0\.0|::)['\"]\s*\)",
        "hardcoded_password": r"(password|passwd|pwd|secret|token)\s*=\s*['\"][^'\"]+['\"]",
        "hardcoded_tmp": r"(/tmp|/var/tmp|C:\\TEMP)(\\\\|/)[^'\"]+",
        "hardcoded_ssl_cert": r"context\.load_cert_chain\(['\"]",
        "aws_keys": r"(AKIA|ASIA)[A-Z0-9]{16}",
        "pickle_used": r"import\s+pickle",
        "marshal_used": r"marshal\.loads\(",
        "md5_used": r"hashlib\.md5\(",
        "cgi_used": r"import\s+cgi",
        "ftplib_used": r"import\s+ftplib",
        "mktemp_used": r"tempfile\.mktemp\(",
        "eval_used": r"eval\(",
        "mark_safe_used": r"mark_safe\(",
        "httpsconnection_used": r"HTTPSConnection\s*\(",
        "urlopen_used": r"urllib\.(request\.)?urlopen\s*\(",
        "random_used": r"import\s+random",
        "telnetlib_used": r"import\s+telnetlib",
        "cElementTree_used": r"from\s+xml\.etree\s+import\s+cElementTree",
        "paramiko_insecure": r"\.set_missing_host_key_policy\(",
        "ssl_bad_version": r"PROTOCOL_(SSLv2|SSLv3|TLSv1|TLSv1\.1)\b",
        "ssl_bad_defaults": r"OP_NO_SSLv2|OP_NO_SSLv3",
        "paramiko_exec_command": r"\.exec_command\s*\(",
        "subprocess_shell_true": r"subprocess\.(Popen|run|call)\(.*shell\s*=\s*True",
        "subprocess_without_shell": r"subprocess\.(Popen|run|call)\(.*shell\s*=\s*False",
        "other_shell_true": r"\b(shell|executable)\s*=\s*True",
        "partial_path_process": r"(Popen|run|call)\(['\"][^/\\][^'\"]+['\"]",
        "shell_process": r"(Popen|run|call)\(.*, shell=True",
        "hardcoded_sql": r"(SELECT|INSERT|UPDATE|DELETE)\s+.*(FROM|INTO|SET|WHERE)",
        "wildcard_injection": r"(rm|ls|chmod|chown)\s+.*[\*\?\[\]]",
        "django_extra_used": r"\.extra\s*\(",
        "django_rawsql_used": r"\.raw\s*\(|RawSQL\s*\(",
        "logging_insecure": r"logging\.basicConfig\(.*handlers\s*=\s*\[.*SocketHandler",
        "jinja2_autoescape_off": r"Environment\s*\(.*autoescape\s*=\s*False",
        "aes_usage": r"AES\.new\(",
        "fernet_usage": r"Fernet\(",
        "xor_usage": r"\bXOR\b",
        "socket_creation": r"socket\.socket\(",
        "server_creation": r"(bind|listen)\s*\(",
        "camera_access": r"(cv2\.VideoCapture|picamera\.PiCamera)",
        "microphone_access": r"pyaudio\.PyAudio",
        "gps_access": r"(gps\.gps|geopy\.Nominatim)",
        "filesystem_ops": r"(open\(|os\.remove\(|shutil\.rmtree\()",
        "shell_injection": r"subprocess\.run\([^)]*shell\s*=\s*True",
        "sql_injection": r"(?i)select\s+.*from\s+.*"
    }
    
    if key in pattern_map:
        pattern = pattern_map[key]
        for idx, line in enumerate(lines, 1):
            if re.search(pattern, line, re.IGNORECASE):
                findings.append((str(idx), line.strip()))
        return findings if findings else [("-", line.strip())]
    
    return [("-", "N/A")]